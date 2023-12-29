import tkinter
import customtkinter as CTk
from PIL import ImageTk,Image
import webbrowser
import pathlib
import openpyxl, xlrd
from openpyxl import Workbook
import sqlite3 
from tkinter import messagebox
import tkinter as tk
from tkinter import ttk
#_____________________________________________________________________

#_____________________________________________________________________

CTk.set_appearance_mode("dark")
CTk.set_default_color_theme("dark-blue")
    

app = CTk.CTk()  
app.geometry("600x440")
app.title("Flash'ipsa")


#Permet de kill la page quand on se connecte

def button_youtube():
    webbrowser.open('https://www.youtube.com')


def button_facebook():
    webbrowser.open('https://www.facebook.com/?locale=fr_FR')
    



def button_function():
    app.destroy()            
    
    CTk.set_appearance_mode("dark")
    CTk.set_default_color_theme("dark-blue")
    

    page2= CTk.CTk()  
    page2.geometry("1280x720")
    page2.title('STATIP\'S')
    
    img12=ImageTk.PhotoImage(Image.open("pattern.png"))
    l12=CTk.CTkLabel(master=page2 ,image=img12)
    l12.pack()

    frame2=CTk.CTkFrame(master=page2, width=620, height=660, corner_radius=15)
    frame2.place(relx=0.5, rely=0.5, anchor=tkinter.CENTER)

    l2=CTk.CTkLabel(master=frame2, text="STATIP\'S",font=CTk.CTkFont(family="KG HAPPY", size=40), text_color="#CD6600",)
    l2.place(x=55, y=25)

    l4 = CTk.CTkLabel(master= frame2, text="Nous récoltons vos données afin de vous proposez les meilleurs statistiques :",font=('Century Gothic',14))
    l4.place(x=55, y=100)

    nom=CTk.CTkEntry(master=frame2, width=220, placeholder_text='Nom', height=35,)
    nom.place(x=300, y=180)
    
    l13 = CTk.CTkLabel(master= frame2, text="Qu'elle est votre Nom ?",font=('Century Gothic',18))
    l13.place(x=75, y=180)
    
    prenom=CTk.CTkEntry(master=frame2, width=220, placeholder_text='Prénom',height=35,)
    prenom.place(x=300, y=230)
   
    l14 = CTk.CTkLabel(master= frame2, text="Qu'elle est votre Prénom ?",font=('Century Gothic',18))
    l14.place(x=70, y=230)
   
    l5 = CTk.CTkLabel(master= frame2, text="Etes-vous un/une ?",font=('Century Gothic',18))
    l5.place(x=75, y=283)
   
    
    genre= CTk.CTkComboBox(master= frame2, values=["Homme", "Femme"], width=220, height=35, button_color="#CD6600")
    genre.place(x=300, y=280)
     
    l6 = CTk.CTkLabel(master= frame2, text="Qu'elle est votre classe ?",font=('Century Gothic',18))
    l6.place(x=75, y=333)
 
    classe= CTk.CTkComboBox(master= frame2, values=["Aéro 1","Aéro 2","Aéro 3","Aéro 4", "Aéro 5"], width=220, height=35, button_color="#CD6600")
    classe.place(x=300, y=330)

    l7 = CTk.CTkLabel(master= frame2, text="La matière à réviser ?",font=('Century Gothic',18))
    l7.place(x=75, y=383)


    matiere= CTk.CTkComboBox(master= frame2, values=["Mathématique","Physique","Informatique","Anglais", "Electronique"], width=220, height=35, button_color="#CD6600")
    matiere.place(x=300, y=380)

    def two_funcs(*funcs):
        def two_funcs(*args, **kwargs):
            for f in funcs:
                f(*args,**kwargs)
        return two_funcs

    file=pathlib.Path("Sondage_flaship's.xlsx")
    if file.exists():
        pass
    else:
        file=Workbook()
        sheet=file.active
        sheet['A1']="Nom"
        sheet['B1']="Prénom"
        sheet['C1']="Genre"
        sheet['D1']="Classe"
        sheet['E1']="Matiere"
        sheet['F1']="âge"
        sheet['G1']="Note"

        file.save("Sondage_flaship's.xlsx")

    def submit():
        nom_exel= nom.get()
        prenom_exel= prenom.get()
        genre_exel= genre.get()
        classe_exel= classe.get()
        matiere_exel= matiere.get()
        age_exel= age.get()
        note_exel= note.get()

        file= openpyxl.load_workbook("Sondage_flaship's.xlsx")
        sheet= file.active
        sheet.cell(column=1, row=sheet.max_row+1, value= nom_exel)
        sheet.cell(column=2, row=sheet.max_row, value= prenom_exel)
        sheet.cell(column=3, row=sheet.max_row, value= genre_exel)
        sheet.cell(column=4, row=sheet.max_row, value= classe_exel)
        sheet.cell(column=5, row=sheet.max_row, value= matiere_exel)
        sheet.cell(column=6, row=sheet.max_row, value= age_exel)
        sheet.cell(column=7, row=sheet.max_row, value= note_exel)

        file.save("Sondage_flaship's.xlsx")

    
    
    def clear():
        nom.delete(0, 100)
        prenom.delete(0,100)


    
    def sliding(value):
        my_label.configure(text=value)   

    def sliding_2(value):
        my_label2.configure(text=value)   

    
    l10= CTk.CTkLabel(master= frame2, text="Qu'elle est votre âge ?", font=("Century Gothic",18))
    l10.place(x=70, y=445)

    l11= CTk.CTkLabel(master= frame2, text="Qu'elle est votre note ?", font=("Century Gothic",18))
    l11.place(x=70, y=505)
    
    age = CTk.CTkSlider(frame2, from_=0, to=100, command=sliding, number_of_steps=100, width=245, height=20, progress_color="#CD6600", button_color='#CD6600')
    age.place(x=280, y=450)

    age.set(0)

    my_label = CTk.CTkLabel(master=frame2, text="", font=("Century Gothic",16))
    my_label.place(x=550, y=445)
    
    note = CTk.CTkSlider(frame2, from_=0, to=20, command=sliding_2, number_of_steps=20, width=245, height=20, progress_color="#CD6600" , button_color="#CD6600" )
    note.place(x=280, y=510)

    note.set(0)

    my_label2 = CTk.CTkLabel(master=frame2, text="", font=("Century Gothic",16))
    my_label2.place(x=550, y=510)

    

    button_sumbit = CTk.CTkButton(master=frame2,text="Ok", font=("Century Gothic",16), command= two_funcs(submit, clear))
    button_sumbit.place(x=380, y=578)

    def boutton_jouer():
 #________________________________________       
        def create_tables(conn):
            cursor = conn.cursor()

            # Create flashcard_sets table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS flashcard_sets (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL
                )
            ''')

            # Create flashcards table with foreign key reference to flashcard_sets
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS flashcards (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    set_id INTEGER NOT NULL,
                    word TEXT NOT NULL,
                    definition TEXT NOT NULL,
                    FOREIGN KEY (set_id) REFERENCES flashcard_sets(id)               
                )
            ''')

        # Add a new flashcard set to the database
        def add_set(conn, name):
            cursor = conn.cursor()

            # Insert the set name into flashcard_sets table
            cursor.execute('''
                INSERT INTO flashcard_sets (name)
                VALUES (?)
            ''', (name,))

            set_id = cursor.lastrowid
            conn.commit()

            return set_id

        # Function to add a flashcard to the database
        def add_card(conn, set_id, word, definition):
            cursor = conn.cursor()

            # Execute SQL query to insert a new flashcard into the database
            cursor.execute('''
                INSERT INTO flashcards (set_id, word, definition)
                VALUES (?, ?, ?)
            ''', (set_id, word, definition))

            # Get the ID of the newly inserted card
            card_id = cursor.lastrowid
            conn.commit()

            return card_id

        # Function to retrieve all flashcard sets from the database
        def get_sets(conn):
            cursor = conn.cursor()

            # Execite SQL query to fetch all flashcard sets
            cursor.execute('''
                SELECT id, name FROM flashcard_sets
            ''')

            rows = cursor.fetchall()
            sets = {row[1]: row[0] for row in rows} # Create a dictionary of sets (name: id)

            return sets

        # Function to retrieve all flashcards of a specific set
        def get_cards(conn, set_id):
            cursor = conn.cursor()

            cursor.execute('''
                SELECT word, definition FROM flashcards
                WHERE set_id = ?
            ''', (set_id,))

            rows = cursor.fetchall()
            cards = [(row[0], row[1]) for row in rows] # Create a list of cards (word, definition)

            return cards

        # Function to delete a flashcard set from the database
        def delete_set(conn, set_id):
            cursor = conn.cursor()

            # Execute SQL query to delete a flashcard set
            cursor.execute('''
                DELETE FROM flashcard_sets
                WHERE id = ?
            ''', (set_id,))

            conn.commit()
            sets_combobox.set('')
            clear_flashcard_display()
            populate_sets_combobox()

            # Clear the current_cards list and reset card_index
            global current_cards, card_index
            current_cards = []
            card_index = 0






        # Function to create a new flashcard set
        def create_set():
            set_name = set_name_var.get()
            if set_name:
                if set_name not in get_sets(conn):
                    set_id = add_set(conn, set_name)
                    populate_sets_combobox()
                    set_name_var.set('')

                    # Clear the input fields
                    set_name_var.set('')
                    word_var.set('')
                    definition_var.set('')

        def add_word():
            set_name = set_name_var.get()
            word = word_var.get()
            definition = definition_var.get()

            if set_name and word and definition:
                if set_name not in get_sets(conn):
                    set_id = add_set(conn, set_name)
                else:
                    set_id = get_sets(conn)[set_name]

                add_card(conn, set_id, word, definition)

                word_var.set('')
                definition_var.set('')

                populate_sets_combobox()

        def populate_sets_combobox():
            sets_combobox['values'] = tuple(get_sets(conn).keys())

        # Function to delete a selected flashcard set
        def delete_selected_set():
            set_name = sets_combobox.get()

            if set_name:
                result = messagebox.askyesno(
                    'Confirmation', f'Are you sure you want to delete the "{set_name}" set?'
                )

                if result == tk.YES:
                    set_id = get_sets(conn)[set_name]
                    delete_set(conn, set_id)
                    populate_sets_combobox()
                    clear_flashcard_display()

        def select_set():
            set_name = sets_combobox.get()

            if set_name:
                set_id = get_sets(conn)[set_name]
                cards = get_cards(conn, set_id)

                if cards:
                    display_flashcards(cards)
                else:
                    word_label.config(text="No cards in this set")
                    definition_label.config(text='')
            else:
                # Clear the current cards list and reset card index
                global current_cards, card_index
                current_cards = []
                card_index = 0
                clear_flashcard_display()

        def display_flashcards(cards):
            global card_index
            global current_cards

            card_index = 0
            current_cards = cards
            
            # Clear the display
            if not cards:
                clear_flashcard_display()
            else:
                show_card()
            
            show_card()

        def clear_flashcard_display():
            word_label.config(text='')
            definition_label.config(text='')

        # Function to display the current flashcards word
        def show_card():
            global card_index
            global current_cards

            if current_cards:
                if 0 <= card_index < len(current_cards):
                    word, _ = current_cards[card_index]
                    word_label.config(text=word)
                    definition_label.config(text='')
                else:
                    clear_flashcard_display()
            else:
                clear_flashcard_display()

        # Function to flip the current card and display its definition
        def flip_card():
            global card_index
            global current_cards

            if current_cards:
                _, definition = current_cards[card_index]
                definition_label.config(text=definition)

        # Function to move to the next card
        def next_card():
            global card_index
            global current_cards

            if current_cards:
                card_index = min(card_index + 1, len(current_cards) -1)
                show_card()

        # Function to move to the previous card
        def prev_card():
            global card_index
            global current_cards

            if current_cards:
                card_index = max(card_index - 1, 0)
                show_card()

        if __name__ == '__main__':
            # Connect to the SQLite database and create tables
            conn = sqlite3.connect('flashcards.db')
            create_tables(conn)

            # Create the main GUI window
            root = CTk.CTk()
            root.title('Flashcards App')
            root.geometry('500x400')

            # Set up variables for storing user input
            set_name_var = tk.StringVar()
            word_var = tk.StringVar()
            definition_var = tk.StringVar()

            # Create a notebook widget to manage tabs
            notebook = ttk.Notebook(root)
            notebook.pack(fill='both', expand=True)

            # Create the "Create Set" tab and its content
            create_set_frame = ttk.Frame(notebook)
            notebook.add(create_set_frame, text='Create Set')

            # Label and Entry widgets for entering set name, word and definition
            tk.Label(create_set_frame, text='Set Name:').pack(padx=5, pady=5)
            ttk.Entry(create_set_frame, textvariable=set_name_var, width=30).pack(padx=5, pady=5)

            ttk.Label(create_set_frame, text='Word:').pack(padx=5, pady=5)
            ttk.Entry(create_set_frame, textvariable=word_var, width=30).pack(padx=5, pady=5)

            ttk.Label(create_set_frame, text='Definition:').pack(padx=5, pady=5)
            ttk.Entry(create_set_frame, textvariable=definition_var, width=30).pack(padx=5, pady=5)

            # Button to add a word to the set 
            ttk.Button(create_set_frame, text='Add Word', command=add_word).pack(padx=5, pady=10)
            
            # Button to save the set 
            ttk.Button(create_set_frame, text='Save Set', command=create_set).pack(padx=5, pady=10)







            # Create the "Select Set" tab and its content
            select_set_frame = ttk.Frame(notebook)
            notebook.add(select_set_frame, text="Select Set")

            # Combobox widget for selecting existing flashcard sets
            sets_combobox = ttk.Combobox(select_set_frame, state='readonly')
            sets_combobox.pack(padx=5, pady=40)

            # Button to select a set 
            ttk.Button(select_set_frame, text='Select Set', command=select_set).pack(padx=5, pady=5)

            # Button to delete a set 
            ttk.Button(select_set_frame, text='Delete Set', command=delete_selected_set).pack(padx=5, pady=5)

            # Create the "Learn mode" tab and its content
            flashcards_frame = ttk.Frame(notebook)
            notebook.add(flashcards_frame, text='Learn Mode')

            # Initialize variables for tracking card index and current cards
            card_index = 0
            current_tabs = []

            # Label to display the word on flashcards
            word_label = ttk.Label(flashcards_frame, text='', font=('TkDefaultFont', 24))
            word_label.pack(padx=5, pady=40)

            # Label to display the definition on flashcards
            definition_label = ttk.Label(flashcards_frame, text='')
            definition_label.pack(padx=5, pady=5)

            # Button to flip the flashcard 
            ttk.Button(flashcards_frame, text='Flip', command=flip_card).pack(side='left', padx=5, pady=5)

            # Button to view the next flashcard 
            ttk.Button(flashcards_frame, text='Next', command=next_card).pack(side='right', padx=5, pady=5)

            # Button to view the previous flashcard 
            ttk.Button(flashcards_frame, text='Previous', command=prev_card).pack(side='right', padx=5, pady=5)

            populate_sets_combobox()

            root.mainloop()


    
    
    
    
    
    
    
    
    
    
    
    
    button_jouer = CTk.CTkButton(master=frame2,text="Jouer", font=("Century Gothic",16), command= boutton_jouer)
    button_jouer.place(x=180, y=578)
    
    
    page2.mainloop()
    












img1=ImageTk.PhotoImage(Image.open("pattern.png"))
l1=CTk.CTkLabel(master=app,image=img1)
l1.pack()

# On cree la fenetre custom dans la fenetre
frame=CTk.CTkFrame(master=l1, width=320, height=360, corner_radius=15)
frame.place(relx=0.5, rely=0.5, anchor=tkinter.CENTER)

l2=CTk.CTkLabel(master=frame, text="Flash'ips",font=CTk.CTkFont(family="KG HAPPY", size=40), text_color="#CD6600",)
l2.place(x=55, y=25)

entry1=CTk.CTkEntry(master=frame, width=220, placeholder_text='Identifiant')
entry1.place(x=50, y=110)

entry2=CTk.CTkEntry(master=frame, width=220, placeholder_text='Mot de passe', show="*")
entry2.place(x=50, y=165)

l3=CTk.CTkLabel(master=frame, text="Forget password?",font=('Century Gothic',12))
l3.place(x=155,y=195)

#Creation des boutons Custom et leurs actions
button1 = CTk.CTkButton(master=frame, width=220, text="Login", command=button_function, corner_radius=6)
button1.place(x=50, y=240)


img2= CTk.CTkImage(Image.open("youtube.png").resize((20,20)))
img3= CTk.CTkImage(Image.open("facebook.png").resize((20,20)))
button2= CTk.CTkButton(master=frame, image=img2, text="Youtube", width=100, height=20, compound="left", fg_color='#2E2E2E', text_color='white', hover_color='#AFAFAF', command=button_youtube)
button2.place(x=50, y=290)

button3= CTk.CTkButton(master=frame, image=img3, text="Facebook", width=100, height=20, compound="left", fg_color='#2E2E2E', text_color='white', hover_color='#AFAFAF', command=button_facebook)
button3.place(x=170, y=290)



 

app.mainloop()
